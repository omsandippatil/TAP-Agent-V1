import io
import os
import re
import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as XLImage

from app.pipeline.llm import LLM_UNAVAILABLE_EVIDENCE

ARIAL = "Arial"
HDR_FILL = PatternFill("solid", fgColor="0F3D3E")
SUBHDR_FILL = PatternFill("solid", fgColor="146B65")
SUB_FILL = PatternFill("solid", fgColor="E6F5F3")
HIGHLIGHT_FILL = PatternFill("solid", fgColor="F5C518")
FLAG_FILL_HIGH = PatternFill("solid", fgColor="FBEAEA")
FLAG_FILL_MEDIUM = PatternFill("solid", fgColor="FEF7DC")
UNAVAILABLE_FILL = PatternFill("solid", fgColor="FBEAEA")
LOW_CONF_FILL = PatternFill("solid", fgColor="FEF7DC")
BANNER_FILL = PatternFill("solid", fgColor="0F3D3E")
WHITE_BOLD = Font(name=ARIAL, bold=True, color="FFFFFF", size=11)
BANNER_TITLE = Font(name=ARIAL, bold=True, color="F5C518", size=14)
BANNER_SUB = Font(name=ARIAL, color="B8E3E0", size=9)
TEAL_BOLD = Font(name=ARIAL, bold=True, color="0F3D3E", size=10)
YELLOW_BOLD = Font(name=ARIAL, bold=True, color="8A6200", size=10)
BOLD = Font(name=ARIAL, bold=True, size=10)
BODY = Font(name=ARIAL, size=10)
SMALL = Font(name=ARIAL, size=9, color="666666")
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")
LEFT_TOP = Alignment(wrap_text=True, vertical="top", horizontal="left")
THIN = Border(bottom=Side(style="thin", color="DDDDDD"))

_HIGHLIGHT_PATTERN = re.compile(r"\*\*(.+?)\*\*")

LOGO_PATH = os.path.join(os.path.dirname(__file__), "..", "render", "assets", "ff_logo.png")

SOURCE_LABELS = {
    "india_csr_page": "Company India CSR page",
    "mca_portal": "MCA portal (verified)",
    "mca_via_search": "MCA via web search (proxy)",
    "national_csr_portal": "National CSR Portal",
    "annual_report": "Annual / sustainability report",
    "global_annual_report": "Annual / sustainability report",
    "partner_search": "Partner-focused web search",
    "people_search": "LinkedIn / people search",
    "plans_search": "Partnerships & plans search",
    "sector_eligibility_search": "Sector & eligibility search",
}


def _strip_highlight_markers(text: str) -> str:
    return _HIGHLIGHT_PATTERN.sub(r"\1", text or "")


def _sheet(wb, title, widths):
    ws = wb.create_sheet(title)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    last_col = get_column_letter(len(widths))
    ws.print_area = f"A1:{last_col}200"
    return ws


def _banner(ws, title_text, subtitle_text, col_span):
    ws.row_dimensions[1].height = 34
    ws.row_dimensions[2].height = 16
    last_col = get_column_letter(col_span)
    ws.merge_cells(f"A1:{last_col}1")
    ws.merge_cells(f"A2:{last_col}2")
    for row in (1, 2):
        for col in range(1, col_span + 1):
            ws.cell(row=row, column=col).fill = BANNER_FILL
    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.font = BANNER_TITLE
    title_cell.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    subtitle_cell = ws.cell(row=2, column=1, value=subtitle_text)
    subtitle_cell.font = BANNER_SUB
    subtitle_cell.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    if os.path.exists(LOGO_PATH):
        img = XLImage(LOGO_PATH)
        img.height = 42
        img.width = 42
        img.anchor = f"{last_col}1"
        ws.add_image(img)
    return 4


def _header(ws, row, texts, col_widths=None):
    ws.row_dimensions[row].height = 30
    for col, t in enumerate(texts, 1):
        c = ws.cell(row=row, column=col, value=t)
        c.font, c.fill, c.alignment = WHITE_BOLD, HDR_FILL, WRAP_CENTER
    return row + 1


def _row(ws, row, values, fonts=None, fill=None, highlight_col_indices=None, alignments=None):
    highlight_col_indices = highlight_col_indices or set()
    for col, v in enumerate(values, 1):
        display_value = _strip_highlight_markers(v) if isinstance(v, str) else v
        c = ws.cell(row=row, column=col, value=display_value)
        c.font = (fonts or [BODY] * len(values))[col - 1]
        c.alignment = (alignments or [WRAP] * len(values))[col - 1]
        c.border = THIN
        if fill:
            c.fill = fill
        elif isinstance(v, str) and v == LLM_UNAVAILABLE_EVIDENCE:
            c.fill = UNAVAILABLE_FILL
        elif col - 1 in highlight_col_indices and isinstance(v, str) and _HIGHLIGHT_PATTERN.search(v):
            match = _HIGHLIGHT_PATTERN.search(v)
            c.comment = Comment(f"Highlighted in report: {match.group(1)}", "FundFinder")
    return row + 1


def _source_hierarchy_note(source_name: str) -> str:
    hierarchy = {
        "india_csr_page": "1 — company's own filings (the spine)",
        "annual_report": "1 — company's own filings (the spine)",
        "global_annual_report": "1 — company's own filings",
        "mca_portal": "2 — statutory regulator filing (CSR-2 by CIN)",
        "national_csr_portal": "2 — statutory regulator data",
        "partner_search": "3 — partner and programme pages",
        "people_search": "3 — decision-maker / people search",
        "plans_search": "3 — partnership and announced-plans search",
        "sector_eligibility_search": "3 — sector and eligibility signals",
        "mca_via_search": "4 — proxy, verify against the actual filing",
        "web_search_snippet": "4 — cross-check only, never the primary figure",
    }
    return hierarchy.get(source_name, "")


async def generate_deep_dive_xlsx(company: str, result: dict, cfg: dict) -> bytes:
    analysis = result.get("analysis") or {}
    breakdown = result.get("score_breakdown", {}) or {}
    sources = result.get("sources", []) or []
    decision_makers = result.get("decision_makers", []) or []
    important_links = result.get("important_links", []) or []
    fit_score = result.get("fit_score", 0)
    tier = result.get("scoring_tier", {}) or {}
    state = result.get("state", "")
    now = datetime.datetime.now().strftime("%d %B %Y")
    has_analysis = bool(analysis)

    wb = Workbook()
    wb.remove(wb.active)

    ws = _sheet(wb, "1. Verdict", [26, 95])
    r = _banner(ws, f"{company} — DEEP-DIVE BASE",
                f"Generated {now}  ·  engine draft — every figure must be verified by a person before this ships",
                2)

    delivery_model = analysis.get("delivery_model", "UNCLEAR") if has_analysis else "UNCLEAR"
    way_in = {
        "FUNDER": "Enter as a grantee — they already fund NGO implementation partners.",
        "HYBRID": "Enter as a delivery-excellence partner strengthening their education portfolio.",
        "IMPLEMENTER": "Enter as a specialist curriculum/tech partner, not a grant recipient.",
    }.get(delivery_model, "To confirm — delivery model unclear from sources.")

    if has_analysis:
        criteria = analysis.get("criteria", []) or []
        weakest = sorted(criteria, key=lambda x: x["score"])[:2]
        red_flags = analysis.get("red_flags", []) or []
        catch = (
            "; ".join(f"{f.get('flag', '')} ({f.get('severity', '')})" for f in red_flags[:2])
            if red_flags
            else "Weakest criteria: " + "; ".join(f"{w['name']} ({w['score']}/5)" for w in weakest)
        )
        call_line = f"{tier.get('label', '')} — fit score {fit_score}/100 — semantic alignment {analysis.get('overall_semantic_alignment', 0)}/100"
        csr_head_note = analysis.get("csr_head_note", "")
        authenticity_note = f"Evidence authenticity {analysis.get('overall_authenticity_score', 0)}/100 · avg criteria confidence {breakdown.get('average_confidence_pct', 0)}%"
        fit_rationale = analysis.get("fit_rationale", "")
    else:
        catch = LLM_UNAVAILABLE_EVIDENCE
        call_line = f"Analysis unavailable — fit score {fit_score}/100 (state: {state})"
        csr_head_note = LLM_UNAVAILABLE_EVIDENCE
        authenticity_note = LLM_UNAVAILABLE_EVIDENCE
        fit_rationale = LLM_UNAVAILABLE_EVIDENCE

    for label, value in [
        ("The call", call_line),
        ("Why it lands there", result.get("strategic_insight", "")),
        ("Fit rationale", fit_rationale),
        ("The catch", catch),
        ("The way in", way_in),
        ("Evidence confidence", authenticity_note),
        ("CSR head philosophy", csr_head_note),
    ]:
        ws.row_dimensions[r].height = 34
        label_cell = ws.cell(row=r, column=1, value=label)
        label_cell.font, label_cell.alignment, label_cell.fill = TEAL_BOLD, LEFT_TOP, SUB_FILL
        value_cell = ws.cell(row=r, column=2, value=_strip_highlight_markers(value))
        value_cell.font, value_cell.alignment = BODY, WRAP
        if isinstance(value, str) and _HIGHLIGHT_PATTERN.search(value):
            match = _HIGHLIGHT_PATTERN.search(value)
            value_cell.comment = Comment(f"Highlighted in report: {match.group(1)}", "FundFinder")
        r += 1
    ws.freeze_panes = "A5"

    ws = _sheet(wb, "2. Fit Against Criteria", [38, 9, 12, 62, 52])
    r = _banner(ws, "Fit Against Criteria", "Every criterion scored 0–5 with confidence and cited evidence", 5)
    r = _header(ws, r, ["Criterion (0–5)", "Score", "Confidence", "Evidence", "Reasoning"])
    if has_analysis:
        for cri in analysis.get("criteria", []):
            row_start = r
            r = _row(ws, r, [
                cri["name"], cri["score"], f"{cri.get('confidence', 0)}%",
                cri.get("evidence", ""), cri.get("reasoning", ""),
            ], alignments=[LEFT_TOP, WRAP_CENTER, WRAP_CENTER, WRAP, WRAP], highlight_col_indices={3})
            if cri.get("confidence", 100) < 50:
                ws.cell(row=row_start, column=3).fill = LOW_CONF_FILL
        avg_conf = breakdown.get("average_confidence_pct", "")
        r = _row(
            ws, r,
            ["OVERALL SEMANTIC ALIGNMENT", analysis.get("overall_semantic_alignment", 0),
             f"avg conf {avg_conf}%", analysis.get("alignment_rationale", ""), ""],
            fonts=[TEAL_BOLD, TEAL_BOLD, TEAL_BOLD, SMALL, SMALL],
            alignments=[LEFT_TOP, WRAP_CENTER, WRAP_CENTER, WRAP, WRAP], fill=SUB_FILL,
        )
        _row(ws, r, ["Evidence authenticity (overall)", analysis.get("overall_authenticity_score", 0),
                     "—", analysis.get("source_quality_assessment", ""), ""],
             fonts=[TEAL_BOLD, BODY, BODY, SMALL, SMALL],
             alignments=[LEFT_TOP, WRAP_CENTER, WRAP_CENTER, WRAP, WRAP])
    else:
        _row(ws, r, ["All criteria", "—", "—", LLM_UNAVAILABLE_EVIDENCE, ""], fill=UNAVAILABLE_FILL)
    ws.freeze_panes = "A5"

    ws = _sheet(wb, "3. Financials & Geography", [30, 24, 20, 72])
    r = _banner(ws, "Financials & Geography", "CSR spend history, mandate status, and geographic footprint", 4)
    r = _header(ws, r, ["Figure / fact", "Value", "Source / confidence", "Evidence excerpt"])
    if has_analysis:
        spend = analysis.get("spend", {}) or {}
        spend_display = spend.get("display") or "Not publicly disclosed"
        r = _row(ws, r, [
            "Latest-year India CSR spend",
            f"{spend_display} ({spend.get('fiscal_year', '')})".strip(),
            f"confidence {spend.get('confidence', 0)}%",
            spend.get("source_excerpt", ""),
        ], alignments=[LEFT_TOP, WRAP, WRAP_CENTER, WRAP])
        if spend.get("trend_direction") and spend.get("trend_direction") != "UNKNOWN":
            r = _row(ws, r, [
                "CSR spend trend", spend.get("trend_direction", ""), "—", spend.get("trend_evidence", ""),
            ], alignments=[LEFT_TOP, WRAP, WRAP_CENTER, WRAP])
        for entry in spend.get("history", []) or []:
            r = _row(ws, r, [
                f"CSR spend — {entry.get('fiscal_year', '')}", entry.get("display", ""),
                "—", entry.get("source_excerpt", ""),
            ], alignments=[LEFT_TOP, WRAP, WRAP_CENTER, WRAP])
        eligibility = analysis.get("eligibility", {}) or {}
        r = _row(ws, r, [
            "Section 135 mandate", eligibility.get("plausibly_mandated", "UNKNOWN"),
            eligibility.get("net_worth_turnover_signal", ""), eligibility.get("reasoning", ""),
        ], alignments=[LEFT_TOP, WRAP, WRAP, WRAP])
        group_foundation = analysis.get("group_foundation", {}) or {}
        if group_foundation.get("routed_through_group"):
            r = _row(ws, r, [
                "Group foundation routing",
                group_foundation.get("foundation_name", "") or "Named group foundation",
                "—", group_foundation.get("explanation", ""),
            ], alignments=[LEFT_TOP, WRAP, WRAP_CENTER, WRAP], fill=FLAG_FILL_MEDIUM)
        for geography in analysis.get("geographies", []) or []:
            r = _row(ws, r, [
                f"Geography: {geography.get('place', '')}", "Mentioned in evidence",
                "—", geography.get("source_excerpt", ""),
            ], alignments=[LEFT_TOP, WRAP, WRAP_CENTER, WRAP])
    else:
        r = _row(ws, r, ["All figures", LLM_UNAVAILABLE_EVIDENCE, "—", ""], fill=UNAVAILABLE_FILL)
    ws.row_dimensions[r + 1].height = 24
    note_cell = ws.cell(
        row=r + 1, column=1,
        value="Every figure above must trace to a primary source. Double-verify before use; no proxy or assumed data.",
    )
    note_cell.font = SMALL
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=4)
    ws.freeze_panes = "A5"

    ws = _sheet(wb, "4. Programmes & Partners", [38, 15, 13, 20, 62])
    r = _banner(ws, "Programmes & Partners", "Named programmes and funded implementation partners", 5)
    r = _header(ws, r, ["Programme / partner", "Type", "Multi-year?", "Scale / relationship", "Evidence"])
    if has_analysis:
        for programme in analysis.get("programmes", []) or []:
            r = _row(ws, r, [
                programme.get("name", ""), "Programme",
                "Yes" if programme.get("is_multi_year") else "One-off / unclear",
                programme.get("cohort_or_scale", ""),
                f"{programme.get('description', '')} — {programme.get('source_excerpt', '')}".strip(" —"),
            ], alignments=[LEFT_TOP, WRAP_CENTER, WRAP_CENTER, WRAP, WRAP])
        for partner in analysis.get("partners", []) or []:
            r = _row(ws, r, [
                partner.get("name", ""), "Partner",
                "—", partner.get("relationship_type", ""), partner.get("source_excerpt", ""),
            ], alignments=[LEFT_TOP, WRAP_CENTER, WRAP_CENTER, WRAP, WRAP])
        if not (analysis.get("programmes") or analysis.get("partners")):
            r = _row(ws, r, ["No programmes or partners found in fetched sources", "", "", "", ""])
    else:
        r = _row(ws, r, ["All programmes/partners", "", "", "", LLM_UNAVAILABLE_EVIDENCE], fill=UNAVAILABLE_FILL)
    ws.freeze_panes = "A5"

    ws = _sheet(wb, "5. Decision-Makers", [26, 32, 18, 56, 30])
    r = _banner(ws, "Decision-Makers", "Named CSR contacts and tenure signals — verify before outreach", 5)
    r = _header(ws, r, ["Name", "Title", "Tenure status", "Evidence", "Source"])
    if decision_makers:
        for person in decision_makers:
            tenure_display = str(person.get("tenure_status", "UNKNOWN")).replace("_", " ").title()
            row_start = r
            r = _row(ws, r, [
                person.get("name", ""), person.get("title", ""), tenure_display,
                person.get("source_excerpt", ""), SOURCE_LABELS.get(person.get("source", ""), person.get("source", "")),
            ], alignments=[LEFT_TOP, WRAP, WRAP_CENTER, WRAP, WRAP])
            if person.get("tenure_status") == "NEW_UNDER_1YR":
                ws.cell(row=row_start, column=3).fill = HIGHLIGHT_FILL
                ws.cell(row=row_start, column=3).font = YELLOW_BOLD
    else:
        r = _row(ws, r, ["No decision-makers found in fetched sources", "", "", "", ""])
    ws.row_dimensions[r + 1].height = 24
    footnote_cell = ws.cell(
        row=r + 1, column=1,
        value="Only current, high-confidence CSR/sustainability roles are listed here — always verify before outreach.",
    )
    footnote_cell.font = SMALL
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=5)
    ws.freeze_panes = "A5"

    ws = _sheet(wb, "6. Governance & Commitment", [28, 14, 78])
    r = _banner(ws, "Governance & Commitment", "Board affinity, volunteering, and open-call signals", 3)
    r = _header(ws, r, ["Signal", "Present", "Detail"])
    if has_analysis:
        board_affinity = analysis.get("board_affinity", {}) or {}
        r = _row(ws, r, [
            "Board / promoter education affinity",
            "Yes" if board_affinity.get("present") else "No",
            f"{board_affinity.get('person_name', '')} — {board_affinity.get('connection', '')}".strip(" —")
            if board_affinity.get("present") else "No named affinity found.",
        ], alignments=[LEFT_TOP, WRAP_CENTER, WRAP])
        volunteering = analysis.get("volunteering", {}) or {}
        r = _row(ws, r, [
            "Employee volunteering / payroll giving",
            "Yes" if volunteering.get("present") else "No",
            f"{volunteering.get('programme_name', '')} — {volunteering.get('description', '')}".strip(" —")
            if volunteering.get("present") else "No named programme found.",
        ], alignments=[LEFT_TOP, WRAP_CENTER, WRAP])
        rfp_signal = analysis.get("rfp_signal", {}) or {}
        r = _row(ws, r, [
            "Open call / RFP for NGO partners",
            "Yes" if rfp_signal.get("present") else "No",
            f"{rfp_signal.get('channel', '')} — {rfp_signal.get('evidence', '')}".strip(" —")
            if rfp_signal.get("present") else "No open call found.",
        ], alignments=[LEFT_TOP, WRAP_CENTER, WRAP], fill=(SUB_FILL if rfp_signal.get("present") else None))
        contact_pathway = analysis.get("contact_pathway", {}) or {}
        r = _row(ws, r, [
            "Contact pathway", "—", contact_pathway.get("channel", "") or "Not identified.",
        ], alignments=[LEFT_TOP, WRAP_CENTER, WRAP], highlight_col_indices={2})
    else:
        r = _row(ws, r, ["All governance signals", "—", LLM_UNAVAILABLE_EVIDENCE], fill=UNAVAILABLE_FILL)
    ws.freeze_panes = "A5"

    ws = _sheet(wb, "7. Red Flags & Open Questions", [98])
    r = _banner(ws, "Red Flags & Open Questions", "Everything the analyst should confirm before outreach", 1)
    r = _header(ws, r, ["Red flags (severity) and open questions to confirm"])
    if has_analysis:
        for flag in analysis.get("red_flags", []) or []:
            r = _row(ws, r, [f"RED FLAG ({flag.get('severity', '')}): {flag.get('flag', '')} — {flag.get('explanation', '')}"],
                      fill=(FLAG_FILL_HIGH if flag.get("severity") == "high" else FLAG_FILL_MEDIUM))
        for question in analysis.get("open_questions", []) or []:
            r = _row(ws, r, [f"To confirm: {question}"])
        if not (analysis.get("red_flags") or analysis.get("open_questions")):
            r = _row(ws, r, ["No red flags or open questions returned."])
    else:
        r = _row(ws, r, [LLM_UNAVAILABLE_EVIDENCE])
    ws.freeze_panes = "A5"

    ws = _sheet(wb, "8. Approach", [26, 100])
    r = _banner(ws, "Approach", "The recommended entry angle given the shape of the portfolio", 2)
    r = _header(ws, r, ["How to enter", "Given the shape of the portfolio"])
    r = _row(ws, r, ["Recommended action", tier.get("action", "")])
    r = _row(ws, r, ["Delivery-model angle", way_in])
    r = _row(ws, r, ["Strategic insight", result.get("strategic_insight", "")], highlight_col_indices={1})
    ws.freeze_panes = "A5"

    ws = _sheet(wb, "9. Sources & Links", [32, 16, 52, 26])
    r = _banner(ws, "Sources & Links", "Every source fetched for this specific company run", 4)
    r = _header(ws, r, ["Source", "Status", "URL", "Hierarchy note"])
    for source in sources:
        if source.get("status") == "NOT_TRIED":
            continue
        row_start = r
        r = _row(ws, r, [
            SOURCE_LABELS.get(source.get("source_name", ""), source.get("source_name", "")),
            source.get("status", ""),
            source.get("url", ""), _source_hierarchy_note(source.get("source_name", "")),
        ], alignments=[LEFT_TOP, WRAP_CENTER, WRAP, WRAP])
        if source.get("status") == "FOUND":
            ws.cell(row=row_start, column=2).fill = SUB_FILL
            ws.cell(row=row_start, column=2).font = TEAL_BOLD
    if important_links:
        r += 1
        r = _header(ws, r, ["Important link (triaged)", "Relevance", "URL", ""])
        for link in important_links:
            r = _row(ws, r, [link.get("label", ""), link.get("relevance", ""), link.get("url", ""), ""],
                      alignments=[LEFT_TOP, WRAP, WRAP, WRAP])
    ws.row_dimensions[r + 1].height = 24
    tail_cell = ws.cell(
        row=r + 1, column=1,
        value="Every source listed was fetched specifically for this company — cross-check the URL domain matches this company before citing.",
    )
    tail_cell.font = SMALL
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=4)
    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


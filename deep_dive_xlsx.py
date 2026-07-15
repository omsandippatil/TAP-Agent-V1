# deep_dive_xlsx.py — 7-sheet deep-dive base per the TAP CSR Research
# Methodology template: Verdict, Fit Against Criteria, Conclusive Evidence,
# Project Ledger, Approach, Open Questions, Sources.
# Gaps are written as "To confirm" with where to confirm them — never filled.
import io
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from methodology import derive_criteria, TO_CONFIRM

ARIAL       = "Arial"
HDR_FILL    = PatternFill("solid", fgColor="4B2E83")   # TAP purple
SUB_FILL    = PatternFill("solid", fgColor="EDE9F8")
CONFIRM_FILL = PatternFill("solid", fgColor="FFF3CD")  # amber for gaps
WHITE_BOLD  = Font(name=ARIAL, bold=True, color="FFFFFF", size=11)
BOLD        = Font(name=ARIAL, bold=True, size=10)
BODY        = Font(name=ARIAL, size=10)
SMALL       = Font(name=ARIAL, size=9, color="666666")
WRAP        = Alignment(wrap_text=True, vertical="top")
THIN        = Border(bottom=Side(style="thin", color="DDDDDD"))


def _sheet(wb, title, widths):
    ws = wb.create_sheet(title)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws


def _header(ws, row, texts):
    for col, t in enumerate(texts, 1):
        c = ws.cell(row=row, column=col, value=t)
        c.font, c.fill, c.alignment = WHITE_BOLD, HDR_FILL, WRAP
    return row + 1


def _row(ws, row, values, fonts=None, fill=None):
    for col, v in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.font = (fonts or [BODY] * len(values))[col - 1]
        c.alignment = WRAP
        c.border = THIN
        if fill:
            c.fill = fill
        elif isinstance(v, str) and v.startswith(TO_CONFIRM):
            c.fill = CONFIRM_FILL
    return row + 1


def generate_deep_dive_xlsx(company: str, result: dict, cfg: dict) -> bytes:
    meth    = derive_criteria(company, result, cfg)
    parsed  = result.get("data", {}) or {}
    bd      = result.get("breakdown", {}) or {}
    sources = result.get("sources", []) or []
    tier    = meth["tier"]
    now     = datetime.datetime.now().strftime("%d %B %Y")

    wb = Workbook()
    wb.remove(wb.active)

    # ── 1. Verdict ───────────────────────────────────────────────────────────
    ws = _sheet(wb, "1. Verdict", [26, 95])
    r = 1
    c = ws.cell(row=r, column=1, value=f"{company} — DEEP-DIVE BASE")
    c.font = Font(name=ARIAL, bold=True, size=14); r += 1
    ws.cell(row=r, column=1, value=f"Generated {now} · engine draft — every "
            f"figure must be verified by a person before this ships").font = SMALL
    r += 2
    penalties = bd.get("penalties", []) or []
    weak = sorted(meth["criteria"], key=lambda x: x["score"])[:2]
    catch = ("; ".join(p.get("reason", "") for p in penalties[:2])
             if penalties else
             "Weakest criteria: " + "; ".join(f"{w['name']} ({w['score']}/5)"
                                               for w in weak))
    dm = (parsed.get("csr_delivery_model") or {}).get("model", "UNCLEAR")
    way_in = {
        "FUNDER": "Enter as a grantee — they already fund NGO implementation partners.",
        "HYBRID": "Enter as a delivery-excellence partner strengthening their education portfolio.",
        "IMPLEMENTER": "Enter as a specialist curriculum/tech partner, not a grant recipient.",
    }.get(dm, f"{TO_CONFIRM} — delivery model unclear from sources.")
    for label, value in [
        ("The call",       f"{tier['label']} — {meth['average']} / 5"
                            f"   (engine fit score {result.get('fit_score', 0)}/100)"),
        ("Why it lands there", result.get("strategic_insight", "")),
        ("The catch",      catch),
        ("The way in",     way_in),
        ("CSR head philosophy", meth["csr_head_note"]),
    ]:
        ws.cell(row=r, column=1, value=label).font = BOLD
        cell = ws.cell(row=r, column=2, value=value)
        cell.font, cell.alignment = BODY, WRAP
        r += 1

    # ── 2. Fit Against Criteria ──────────────────────────────────────────────
    ws = _sheet(wb, "2. Fit Against Criteria", [42, 12, 16, 80])
    r = _header(ws, 1, ["Criterion (0–5)", "Score", "Rating", "Evidence (one line)"])
    for cri in meth["criteria"]:
        r = _row(ws, r, [cri["name"], cri["score"], cri["rating"], cri["evidence"]])
    r = _row(ws, r, ["AVERAGE (of the eight)", meth["average"],
                     tier["label"], "The overall is the average of the eight."],
             fonts=[BOLD, BOLD, BOLD, SMALL], fill=SUB_FILL)
    _row(ws, r, ["CSR head philosophy", "not scored", "—", meth["csr_head_note"]],
         fonts=[BOLD, BODY, BODY, SMALL])

    # ── 3. Conclusive Evidence ───────────────────────────────────────────────
    ws = _sheet(wb, "3. Conclusive Evidence", [30, 26, 22, 80])
    r = _header(ws, 1, ["Figure / fact", "Value", "Source", "Verbatim excerpt"])
    spend = parsed.get("spend") or {}
    r = _row(ws, r, ["Latest-year India CSR spend",
                     spend.get("display") or f"{TO_CONFIRM} — CSR-2 via CIN",
                     spend.get("source_type", "—"), spend.get("excerpt", "")])
    for f in (parsed.get("focus_areas") or [])[:8]:
        r = _row(ws, r, [f"CSR focus: {f.get('value','')}",
                         f.get("confidence", ""), f.get("source_type", ""),
                         f.get("excerpt", "")])
    for cl in (bd.get("adjacency_boost", {}).get("fired_clusters") or [])[:5]:
        exc = (cl.get("evidence_excerpts") or [""])[0]
        r = _row(ws, r, [f"Adjacency: {cl.get('label','')}",
                         ", ".join(cl.get("keywords_found", [])[:3]),
                         "fetched sources", exc])
    ws.cell(row=r + 1, column=1,
            value="Every figure above must trace to a primary source. "
                  "Double-verify before use; no proxy or assumed data.").font = SMALL

    # ── 4. Project / Portfolio Ledger ────────────────────────────────────────
    ws = _sheet(wb, "4. Project Ledger", [45, 24, 85])
    r = _header(ws, 1, ["Project / programme (latest year)", "Source", "Excerpt"])
    progs = parsed.get("programs") or []
    if progs:
        for p in progs:
            r = _row(ws, r, [p.get("name", ""), p.get("source_type", ""),
                             p.get("excerpt", "")])
    else:
        r = _row(ws, r, [f"{TO_CONFIRM} — latest-year project list not found; "
                         f"read from CSR-2 annexure / annual report CSR section",
                         "", ""])
    partners = parsed.get("ngo_partners") or []
    if partners:
        r += 1
        r = _header(ws, r, ["Funded / implementing partner", "TAP-similar?", "Excerpt"])
        for p in partners[:10]:
            sim = ("PEER NGO" if p.get("is_peer_ngo")
                   else "Similar" if p.get("tap_similar") else "—")
            r = _row(ws, r, [p.get("name", ""), sim, p.get("excerpt", "")])

    # ── 5. Approach ──────────────────────────────────────────────────────────
    ws = _sheet(wb, "5. Approach", [26, 100])
    r = _header(ws, 1, ["How to enter", "Given the shape of the portfolio"])
    scoring_tier = result.get("scoring_tier", {}) or {}
    r = _row(ws, r, ["Recommended action", scoring_tier.get("action", "")])
    r = _row(ws, r, ["Delivery-model angle", way_in])
    r = _row(ws, r, ["Strategic insight", result.get("strategic_insight", "")])

    # ── 6. Open Questions ────────────────────────────────────────────────────
    ws = _sheet(wb, "6. Open Questions", [110])
    r = _header(ws, 1, ["What still needs confirming (none should change the "
                        "verdict; all sharpen it)"])
    for q in meth["open_questions"]:
        r = _row(ws, r, [f"{TO_CONFIRM}: {q}"])

    # ── 7. Sources ───────────────────────────────────────────────────────────
    ws = _sheet(wb, "7. Sources", [30, 16, 70, 24])
    r = _header(ws, 1, ["Source", "Status", "URL", "Hierarchy note"])
    hierarchy = {
        "india_csr_page":      "1 — company's own filings (the spine)",
        "annual_report":       "1 — company's own filings (the spine)",
        "global_annual_report":"1 — company's own filings",
        "mca_portal":          "2 — statutory regulator filing (CSR-2 by CIN)",
        "national_csr_portal": "2 — statutory regulator data",
        "partner_search":      "3 — partner and programme pages",
        "people_search":       "3 — partner and programme pages",
        "web_search_snippet":  "4 — cross-check only, never the primary figure",
        "mca_via_search":      "4 — proxy, verify against the actual filing",
    }
    for s in sources:
        if s.get("status") == "NOT_TRIED":
            continue
        sn = s.get("source_name", "")
        r = _row(ws, r, [sn, s.get("status", ""), s.get("url", ""),
                         hierarchy.get(sn, "")])
    ws.cell(row=r + 1, column=1,
            value="Latest financial year only — confirm the newest disclosure "
                  "before relying on any figure above.").font = SMALL

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
